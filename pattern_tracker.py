"""
pattern_tracker.py — MaverickPICKS Pattern Watchlist Tracker
=============================================================
Imports picks from the scanner's out_csv into a persistent
watchlist (pattern_watchlist.json), then checks each pick
daily after market close and logs outcomes to an Excel log.

Workflow (run after 3:30pm IST every trading day):
  Step 1 — scan:    python pattern_detector_v2.py --csv_file NIFTY500_MASTER.csv --out_csv todays_picks.csv
  Step 2 — import:  python pattern_tracker.py --import_csv todays_picks.csv
  Step 3 — check:   python pattern_tracker.py --check
  Or combined:      python pattern_tracker.py --import_csv todays_picks.csv --check

Files created/maintained:
  pattern_watchlist.json   — persistent watchlist (source of truth)
  pattern_tracker_log.xlsx — running Excel log, one row per check event
"""

import argparse
import json
import os
import time
import warnings
from datetime import datetime, timedelta, timezone

import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

WATCHLIST_FILE  = "pattern_watchlist.json"
LOG_FILE        = "pattern_tracker_log.xlsx"

# States
ST_WATCHING     = "WATCHING"
ST_BREAKOUT     = "BREAKOUT"
ST_BREAKDOWN    = "BREAKDOWN"
ST_EXPIRED      = "EXPIRED"

# Colors for Excel
CLR = {
    "header_bg":    "1F3864",   # dark navy
    "header_fg":    "FFFFFF",
    "breakout_bg":  "E2EFDA",   # green tint
    "breakdown_bg": "FCE4D6",   # red tint
    "expired_bg":   "F2F2F2",   # grey
    "expiring_bg":  "FFF2CC",   # amber
    "watching_bg":  "DEEAF1",   # blue tint
    "section_bg":   "D6E4F0",
}

IST = timezone(timedelta(hours=5, minutes=30))


# ─────────────────────────────────────────────────────────────────────────────
# WATCHLIST PERSISTENCE
# ─────────────────────────────────────────────────────────────────────────────

def load_watchlist() -> dict:
    if os.path.exists(WATCHLIST_FILE):
        with open(WATCHLIST_FILE, "r") as f:
            return json.load(f)
    return {"picks": {}, "last_updated": None}


def save_watchlist(wl: dict):
    wl["last_updated"] = datetime.now(IST).strftime("%Y-%m-%d %H:%M IST")
    with open(WATCHLIST_FILE, "w") as f:
        json.dump(wl, f, indent=2)


def _pick_key(symbol: str, pattern: str) -> str:
    """Unique key per symbol+pattern combination."""
    return f"{symbol}|{pattern}"


# ─────────────────────────────────────────────────────────────────────────────
# IMPORT FROM SCANNER CSV
# ─────────────────────────────────────────────────────────────────────────────

REQUIRED_COLS = [
    "Symbol", "Pattern", "Breakout_Level", "Stop_Loss",
    "Target_1", "Risk_Reward", "Confidence", "Score",
]

def import_from_csv(csv_path: str) -> tuple:
    """
    Read scanner output CSV and add new picks to watchlist.
    Skips picks already in watchlist (same symbol+pattern).
    Returns (added_count, skipped_count, watchlist).
    """
    if not os.path.exists(csv_path):
        print(f"  [ERROR] CSV not found: {csv_path}")
        return 0, 0, load_watchlist()

    df = pd.read_csv(csv_path)

    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        print(f"  [ERROR] CSV missing columns: {missing}")
        return 0, 0, load_watchlist()

    wl = load_watchlist()
    added = skipped = 0

    for _, row in df.iterrows():
        symbol  = str(row["Symbol"]).strip()
        pattern = str(row["Pattern"]).strip()
        key     = _pick_key(symbol, pattern)

        if key in wl["picks"] and wl["picks"][key]["state"] == ST_WATCHING:
            skipped += 1
            continue

        # Parse expiry from Narrative if present
        expiry_date = None
        narrative   = str(row.get("Narrative", ""))
        if "Valid until" in narrative:
            try:
                part = narrative.split("Valid until")[1].split("(")[0].strip()
                expiry_date = pd.to_datetime(part, dayfirst=True).strftime("%Y-%m-%d")
            except Exception:
                pass
        if expiry_date is None and "EXPIRING SOON" in narrative:
            try:
                part = narrative.split("by")[1].split(".")[0].strip()
                expiry_date = pd.to_datetime(part, dayfirst=True).strftime("%Y-%m-%d")
            except Exception:
                pass

        pick = {
            "symbol":          symbol,
            "pattern":         pattern,
            "state":           ST_WATCHING,
            "date_added":      datetime.now(IST).strftime("%Y-%m-%d"),
            "breakout_level":  float(row.get("Breakout_Level") or 0),
            "stop_loss":       float(row.get("Stop_Loss")      or 0),
            "target_1":        float(row.get("Target_1")       or 0),
            "risk_reward":     float(row.get("Risk_Reward")    or 0),
            "confidence":      str(row.get("Confidence", "")),
            "score":           float(row.get("Score")          or 0),
            "breakout_vol_watch": float(row.get("Breakout_Vol_Watch") or 0),
            "expiry_date":     expiry_date,
            "pole_return_pct": float(row.get("Pole_Return_%")  or 0),
            "gap_to_break_pct": float(row.get("Gap_To_B_%") or
                                      row.get("Gap_To_Break_%") or 0),
            "weekly_confirmed": bool(row.get("Weekly_Confirmed", False)),
            "date_resolved":   None,
            "resolved_price":  None,
            "resolved_volume": None,
            "resolved_note":   None,
            "price_history":   [],   # list of {date, close, volume, gap_pct}
        }

        wl["picks"][key] = pick
        added += 1

    save_watchlist(wl)
    return added, skipped, wl


# ─────────────────────────────────────────────────────────────────────────────
# PRICE FETCH (single stock, last confirmed close)
# ─────────────────────────────────────────────────────────────────────────────

def _last_trading_day() -> datetime:
    now = datetime.now(IST)
    d   = now.date()
    # Weekend rollback
    if d.weekday() == 5: d -= timedelta(days=1)
    elif d.weekday() == 6: d -= timedelta(days=2)
    else:
        mc = now.replace(hour=15, minute=30, second=0, microsecond=0)
        if now < mc:
            d -= timedelta(days=1)
            if d.weekday() == 6: d -= timedelta(days=2)
            elif d.weekday() == 5: d -= timedelta(days=1)
    return datetime(d.year, d.month, d.day)


def fetch_last_close(symbol: str) -> tuple:
    """Returns (close, volume, date_str) or (None, None, None)."""
    anchor = _last_trading_day()
    end    = anchor + timedelta(days=1)
    start  = anchor - timedelta(days=5)
    for attempt in range(3):
        try:
            df = yf.download(
                symbol, start=start, end=end,
                interval="1d", auto_adjust=False,
                progress=False, threads=False,
            )
            if df.empty:
                raise ValueError("Empty")
            # Flatten multi-index if present
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = [c[0] for c in df.columns]
            df = df.dropna(subset=["Close"])
            if df.empty:
                raise ValueError("No close data")
            last = df.iloc[-1]
            date_str = df.index[-1].strftime("%Y-%m-%d")
            return float(last["Close"]), float(last["Volume"]), date_str
        except Exception:
            time.sleep(1)
    return None, None, None


# ─────────────────────────────────────────────────────────────────────────────
# STATE EVALUATION
# ─────────────────────────────────────────────────────────────────────────────

def evaluate_pick(pick: dict, close: float, volume: float,
                  date_str: str) -> dict:
    """
    Given today's close and volume, determine new state.
    Returns updated pick dict.
    """
    pick = pick.copy()
    brk  = pick["breakout_level"]
    stop = pick["stop_loss"]
    exp  = pick.get("expiry_date")
    bvw  = pick.get("breakout_vol_watch", 0)

    # Gap to breakout
    gap_pct = round((brk - close) / close * 100, 2) if close > 0 else 0

    # Append to price history
    pick["price_history"].append({
        "date":    date_str,
        "close":   round(close, 2),
        "volume":  int(volume),
        "gap_pct": gap_pct,
    })

    # ── Check states in priority order ───────────────────────────────────────

    # 1. Breakout: close above breakout level
    #    Volume confirmation preferred but not required to flag — noted in message
    if close >= brk:
        vol_confirmed = volume >= bvw if bvw > 0 else True
        pick["state"]          = ST_BREAKOUT
        pick["date_resolved"]  = date_str
        pick["resolved_price"] = round(close, 2)
        pick["resolved_volume"]= int(volume)
        pick["resolved_note"]  = (
            f"BREAKOUT ✓ — closed ₹{close:.2f} above ₹{brk:.2f}. "
            + (f"Volume {volume:,.0f} ✓ above watch level {bvw:,.0f}."
               if vol_confirmed
               else f"⚠ Volume {volume:,.0f} BELOW watch level {bvw:,.0f} — confirm on chart.")
        )
        return pick

    # 2. Breakdown: close below stop loss
    if close <= stop:
        pick["state"]          = ST_BREAKDOWN
        pick["date_resolved"]  = date_str
        pick["resolved_price"] = round(close, 2)
        pick["resolved_volume"]= int(volume)
        pick["resolved_note"]  = (
            f"BREAKDOWN ✗ — closed ₹{close:.2f} below stop ₹{stop:.2f}. "
            f"Pattern failed. Remove from watchlist."
        )
        return pick

    # 3. Expired: past expiry date with no breakout
    if exp:
        try:
            exp_dt = pd.to_datetime(exp)
            today  = pd.Timestamp(_last_trading_day())
            if today > exp_dt:
                pick["state"]          = ST_EXPIRED
                pick["date_resolved"]  = date_str
                pick["resolved_price"] = round(close, 2)
                pick["resolved_volume"]= int(volume)
                pick["resolved_note"]  = (
                    f"EXPIRED — pattern exceeded Murphy's time limit (deadline was "
                    f"{exp_dt.strftime('%d-%b-%Y')}). "
                    f"No breakout by expiry. Remove from watchlist."
                )
                return pick
        except Exception:
            pass

    # 4. Still watching
    days_tracked = len(pick["price_history"])
    pick["resolved_note"] = (
        f"Watching — day {days_tracked}. "
        f"Close ₹{close:.2f}, {gap_pct:.1f}% below breakout ₹{brk:.2f}."
    )
    return pick


# ─────────────────────────────────────────────────────────────────────────────
# CHECK ALL WATCHING PICKS
# ─────────────────────────────────────────────────────────────────────────────

def check_watchlist() -> dict:
    """
    Fetch latest prices for all WATCHING picks, evaluate state,
    update watchlist, return summary.
    """
    wl      = load_watchlist()
    picks   = wl["picks"]
    today   = datetime.now(IST).strftime("%Y-%m-%d")

    watching  = [k for k, p in picks.items() if p["state"] == ST_WATCHING]
    breakouts = []
    breakdowns= []
    expireds  = []
    still_watching = []

    print(f"\n  Checking {len(watching)} WATCHING pick(s)...\n")

    for key in watching:
        pick   = picks[key]
        symbol = pick["symbol"]
        print(f"  {symbol:<20} {pick['pattern']:<22}", end=" ")

        close, volume, date_str = fetch_last_close(symbol)

        if close is None:
            print("→ [DATA ERROR] skipping")
            still_watching.append(pick)
            continue

        updated = evaluate_pick(pick, close, volume, date_str)
        picks[key] = updated
        state = updated["state"]

        print(f"→ ₹{close:.2f}  {state}")

        if state == ST_BREAKOUT:  breakouts.append(updated)
        elif state == ST_BREAKDOWN: breakdowns.append(updated)
        elif state == ST_EXPIRED:   expireds.append(updated)
        else:                        still_watching.append(updated)

    save_watchlist(wl)

    return {
        "date":          today,
        "breakouts":     breakouts,
        "breakdowns":    breakdowns,
        "expireds":      expireds,
        "watching":      still_watching,
        "all_picks":     picks,
    }


# ─────────────────────────────────────────────────────────────────────────────
# TERMINAL REPORT
# ─────────────────────────────────────────────────────────────────────────────

def print_report(summary: dict):
    SEP  = "═" * 72
    SEP2 = "─" * 72
    D    = summary["date"]
    bko  = summary["breakouts"]
    bkd  = summary["breakdowns"]
    exp  = summary["expireds"]
    wat  = summary["watching"]

    print(f"\n{SEP}")
    print(f"  MaverickPICKS PATTERN TRACKER — {D}")
    print(f"  Breakouts: {len(bko)}  |  Breakdowns: {len(bkd)}  |  "
          f"Expired: {len(exp)}  |  Watching: {len(wat)}")
    print(SEP)

    def _print_picks(picks, label, emoji):
        if not picks:
            return
        print(f"\n  {emoji}  {label.upper()}")
        print(f"  {SEP2}")
        for p in picks:
            exp_str = f"  exp:{p['expiry_date']}" if p.get('expiry_date') else ""
            print(f"  {p['symbol']:<16} {p['pattern']:<22} "
                  f"Score:{p['score']:>5.0f}  {p['confidence']:<7}{exp_str}")
            print(f"    {p.get('resolved_note','')}")

    # Alerts first
    _print_picks(bko,  "🟢 BREAKOUTS — enter on next open if volume confirmed", "🟢")
    _print_picks(bkd,  "🔴 BREAKDOWNS — pattern failed, exit watchlist",        "🔴")
    _print_picks(exp,  "⚫ EXPIRED — time limit passed, remove from watchlist",  "⚫")

    # Expiring soon within watching
    expiring_soon = [p for p in wat
                     if p.get("expiry_date") and
                     (pd.to_datetime(p["expiry_date"]) -
                      pd.Timestamp(_last_trading_day())).days <= 2]
    still = [p for p in wat if p not in expiring_soon]

    _print_picks(expiring_soon, "⚠ EXPIRING SOON — act or remove",  "⚠")

    if still:
        print(f"\n  👁  WATCHING ({len(still)})")
        print(f"  {SEP2}")
        for p in still:
            h = p.get("price_history", [])
            last_h = h[-1] if h else {}
            gap    = last_h.get("gap_pct", p.get("gap_to_break_pct", 0))
            days   = len(h)
            exp_d  = p.get("expiry_date", "—")
            print(f"  {p['symbol']:<16} {p['pattern']:<22} "
                  f"Gap:{gap:>5.1f}%  day:{days:>2}  "
                  f"brk:₹{p['breakout_level']:.2f}  exp:{exp_d}")

    print(f"\n{SEP}\n")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL LOG
# ─────────────────────────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    "Check_Date", "Symbol", "Pattern", "State", "Days_Tracked",
    "Score", "Confidence", "Close", "Volume", "Gap_To_Break_%",
    "Breakout_Level", "Stop_Loss", "Target_1", "Risk_Reward",
    "Breakout_Vol_Watch", "Expiry_Date", "Weekly_Confirmed",
    "Pole_Return_%", "Date_Added", "Date_Resolved", "Note",
]

def _cell_style(ws, row, col, value, bg=None, bold=False, fmt=None, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if bold:
        cell.font = Font(bold=True, name="Arial", size=10)
    else:
        cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if fmt:
        cell.number_format = fmt
    return cell


def _state_color(state: str) -> str:
    return {
        ST_BREAKOUT:  CLR["breakout_bg"],
        ST_BREAKDOWN: CLR["breakdown_bg"],
        ST_EXPIRED:   CLR["expired_bg"],
        ST_WATCHING:  CLR["watching_bg"],
    }.get(state, "FFFFFF")


def append_to_excel(summary: dict):
    """
    Append today's check results to the running Excel log.
    Creates the file with headers if it doesn't exist.
    One row per pick per check date.
    """
    check_date = summary["date"]
    all_picks  = summary["all_picks"]

    # ── Load or create workbook ───────────────────────────────────────────────
    if os.path.exists(LOG_FILE):
        wb = load_workbook(LOG_FILE)
        ws = wb["Tracker Log"] if "Tracker Log" in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tracker Log"

        # Header row
        for col, h in enumerate(EXCEL_HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill   = PatternFill("solid", start_color=CLR["header_bg"])
            c.font   = Font(bold=True, color=CLR["header_fg"], name="Arial", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    # ── Collect rows for this check ───────────────────────────────────────────
    # Build set of (check_date, symbol, pattern) already in sheet to avoid duplicates
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[2]:
            existing.add((str(row[0]), str(row[1]), str(row[2])))

    rows_added = 0
    for key, pick in all_picks.items():
        h      = pick.get("price_history", [])
        today_h= next((x for x in reversed(h) if x["date"] == check_date), None)

        dedup_key = (check_date, pick["symbol"], pick["pattern"])
        if dedup_key in existing:
            continue

        close  = today_h["close"]   if today_h else pick.get("resolved_price") or ""
        volume = today_h["volume"]  if today_h else pick.get("resolved_volume") or ""
        gap    = today_h["gap_pct"] if today_h else ""

        days_tracked = len(h)

        row_data = [
            check_date,
            pick["symbol"],
            pick["pattern"],
            pick["state"],
            days_tracked,
            pick["score"],
            pick["confidence"],
            close,
            volume,
            gap,
            pick["breakout_level"],
            pick["stop_loss"],
            pick["target_1"],
            pick["risk_reward"],
            pick.get("breakout_vol_watch", ""),
            pick.get("expiry_date", ""),
            pick.get("weekly_confirmed", ""),
            pick.get("pole_return_pct", ""),
            pick["date_added"],
            pick.get("date_resolved", ""),
            pick.get("resolved_note", ""),
        ]

        next_row = ws.max_row + 1
        bg_color = _state_color(pick["state"])
        bold_row = pick["state"] in (ST_BREAKOUT, ST_BREAKDOWN)

        for col, val in enumerate(row_data, 1):
            fmt = None
            align = "left"
            if col in (6, 10, 14, 18):      # Score, Gap%, RR, Pole%
                fmt = "0.0"
                align = "right"
            elif col in (7, 11, 12, 13, 15): # Price cols
                fmt = "#,##0.00"
                align = "right"
            elif col == 9:                    # Volume
                fmt = "#,##0"
                align = "right"
            _cell_style(ws, next_row, col, val,
                        bg=bg_color, bold=bold_row, fmt=fmt, align=align)

        ws.row_dimensions[next_row].height = 18
        rows_added += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [12, 16, 22, 12, 8, 6, 9, 10, 12, 10,
              12, 10, 10, 8, 16, 12, 12, 10, 12, 14, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(LOG_FILE)
    print(f"  Excel log updated — {rows_added} row(s) added → {LOG_FILE}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="MaverickPICKS Pattern Tracker — daily watchlist monitor"
    )
    ap.add_argument("--import_csv", type=str, default="",
                    help="Scanner output CSV to import picks from")
    ap.add_argument("--check", action="store_true",
                    help="Check all WATCHING picks against latest prices")
    ap.add_argument("--status", action="store_true",
                    help="Show current watchlist without fetching prices")
    args = ap.parse_args()

    print(f"\n{'═'*60}")
    print("  MaverickPICKS Pattern Tracker")
    print(f"  {datetime.now(IST).strftime('%d-%b-%Y  %H:%M IST')}")
    print(f"{'═'*60}")

    # ── Import ────────────────────────────────────────────────────────────────
    if args.import_csv:
        print(f"\n  Importing picks from: {args.import_csv}")
        added, skipped, wl = import_from_csv(args.import_csv)
        total = len([p for p in wl["picks"].values()
                     if p["state"] == ST_WATCHING])
        print(f"  Added: {added}  |  Skipped (already watching): {skipped}")
        print(f"  Total WATCHING: {total}")

    # ── Status only ───────────────────────────────────────────────────────────
    if args.status:
        wl    = load_watchlist()
        picks = wl["picks"]
        print(f"\n  Watchlist: {len(picks)} total pick(s)")
        by_state = {}
        for p in picks.values():
            by_state.setdefault(p["state"], []).append(p)
        for state, ps in sorted(by_state.items()):
            print(f"\n  {state} ({len(ps)})")
            for p in ps:
                h = p.get("price_history", [])
                last = h[-1] if h else {}
                print(f"    {p['symbol']:<16} {p['pattern']:<22} "
                      f"day:{len(h):>2}  "
                      f"brk:₹{p['breakout_level']:.2f}  "
                      f"exp:{p.get('expiry_date','—')}")
        return

    # ── Check ─────────────────────────────────────────────────────────────────
    if args.check:
        summary = check_watchlist()
        print_report(summary)
        append_to_excel(summary)

    if not args.import_csv and not args.check and not args.status:
        ap.print_help()


if __name__ == "__main__":
    main()
